# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="C:\CODE\dataGeneration\codes\syndatagenerator\codes\datagenerator\datagenerator\Data.xlsx"
# load demo.xlsx 
wb=load_workbook(filepath)
# select demo.xlsx
sheet=wb.active
# get max row count
max_row=sheet.max_row
# get max column count
max_column=sheet.max_column
# iterate over all cells 
# iterate over all rows
for i in range(1,max_row+1):
     
     # iterate over all columns
     for j in range(1,max_column+1):
          # get particular cell value    
          cell_obj=wb.cell(row=i,column=j)
          # print cell value     
          print(cell_obj.value,end=' | ')
     # print new line
     print('\n')